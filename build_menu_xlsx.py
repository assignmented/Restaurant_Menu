# -*- coding: utf-8 -*-
"""Parse The Black Perch menu pages (menu.tronsart.com) into an Excel workbook."""
import re
import html
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# top-level category + sub-category label per page slug.
# Hub pages (bar, main_course, coffee_shop) only link onward — their leaf
# children are listed directly below.
PAGES = [
    ("breakfast",         "Breakfast",                       "Breakfast"),
    ("appetizers",        "Appetizers & Soups",              "Appetizers & Soups"),
    # Main Course hub -> 7 leaf sub-pages
    ("signature",         "Main Course",                     "Signature Course"),
    ("entrees",           "Main Course",                     "Entrees"),
    ("sea_food",          "Main Course",                     "Sea Food"),
    ("indian_cuisine",    "Main Course",                     "Indian Cuisine"),
    ("swahili_dishes",    "Main Course",                     "Swahili Dishes"),
    ("platters",          "Main Course",                     "Platters"),
    ("side_dishes",       "Main Course",                     "Side Dishes"),
    ("pizzas",            "Pizza, Burgers & Sandwiches",     "Pizza, Burgers & Sandwiches"),
    # Coffee Shop & Ice Cream hub -> 4 leaf sub-pages
    ("coffee",            "Coffee Shop & Ice Cream",         "Coffee Menu"),
    ("icecream",          "Coffee Shop & Ice Cream",         "Ice Cream"),
    ("mojitos",           "Coffee Shop & Ice Cream",         "Mojito & Coladas"),
    ("frappes",           "Coffee Shop & Ice Cream",         "Frappes, Boba & Slushies"),
    ("soft_drink",        "Soft Drinks",                     "Soft Drinks"),
    # Happy Hour hub -> 12 leaf sub-pages
    ("cocktails",         "Happy Hour",                      "Cocktails"),
    ("cocktails_towers",  "Happy Hour",                      "Cocktail Towers"),
    ("wines",             "Happy Hour",                      "Wines"),
    ("beers",             "Happy Hour",                      "Beers"),
    ("beers_cans",        "Happy Hour",                      "Beer Cans"),
    ("vodka",             "Happy Hour",                      "Gin / Vodka"),
    ("whiskey",           "Happy Hour",                      "Whiskey"),
    ("cognac",            "Happy Hour",                      "Cognac"),
    ("brandy",            "Happy Hour",                      "Brandy | Rum"),
    ("liquor",            "Happy Hour",                      "Liquor"),
    ("tequila",           "Happy Hour",                      "Tequila"),
    ("tots",              "Happy Hour",                      "Tots"),
]

TIT2_RE = re.compile(r'<span class="tit2 t-center">\s*(.*?)\s*</span>', re.S)
# a blo3 block: capture name (txt21), desc (txt23), price (txt22)
BLO3_RE = re.compile(
    r'class="txt21 m-b-3">\s*(.*?)\s*</a>.*?'
    r'class="txt23">\s*(.*?)\s*</span>.*?'
    r'class="txt22 m-t-20">\s*(.*?)\s*</span>',
    re.S,
)


def clean(s):
    s = html.unescape(s)
    s = re.sub(r'\s+', ' ', s).strip()
    s = s.replace('\xa0', ' ').strip()
    return s


def price_to_number(p):
    m = re.search(r'([\d,]+)', p)
    if not m:
        return None
    try:
        return int(m.group(1).replace(',', ''))
    except ValueError:
        return None


def parse_page(path, category, subcat):
    """Yield rows (category, subcat, brand, item, desc, price_text, price_num)."""
    with open(path, encoding='utf-8', errors='replace') as f:
        content = f.read()

    # The Happy Hour hub page (bar) links out to sub-pages; it has no blo3 items.
    rows = []
    # Walk the document: track current tit2 section as 'brand'.
    current_brand = subcat  # default if no tit2 section yet
    # Find positions of all tit2 headers
    headers = [(m.start(), m.end(), clean(m.group(1))) for m in TIT2_RE.finditer(content)]

    # If there are tit2 sections, split content by them.
    if headers:
        for i, (s, e, label) in enumerate(headers):
            chunk_start = e
            chunk_end = headers[i + 1][0] if i + 1 < len(headers) else len(content)
            chunk = content[chunk_start:chunk_end]
            # label like "Sweet Red Wine"; if it's just repeating the page banner, keep subcat
            brand = label if label else subcat
            for m in BLO3_RE.finditer(chunk):
                name = clean(m.group(1))
                desc = clean(m.group(2))
                price_text = clean(m.group(3))
                if not name:
                    continue
                rows.append((category, subcat, brand, name, desc, price_text,
                             price_to_number(price_text)))
    else:
        # no tit2 sections found; just grab all blo3
        for m in BLO3_RE.finditer(content):
            name = clean(m.group(1))
            desc = clean(m.group(2))
            price_text = clean(m.group(3))
            if not name:
                continue
            rows.append((category, subcat, current_brand, name, desc, price_text,
                         price_to_number(price_text)))
    return rows


def main():
    all_rows = []
    summary = []  # (category, subcat, item_count)
    for slug, category, subcat in PAGES:
        path = f"page_{slug}.html"
        if not os.path.exists(path):
            print(f"MISSING: {path}")
            continue
        rows = parse_page(path, category, subcat)
        if not rows:
            print(f"  WARN: no items parsed from {path}")
        all_rows.extend(rows)
        summary.append((category, subcat, len(rows)))
        print(f"  {slug:20s} -> {len(rows):3d} items")

    # ---- Build workbook ----
    wb = Workbook()

    # Sheet 1: All Menu Items
    ws = wb.active
    ws.title = "Menu Items"
    headers = ["Category", "Sub-Category", "Brand / Section", "Menu Item",
               "Description / Size", "Price (KSh)", "Price (numeric)"]
    ws.append(headers)
    for r in all_rows:
        ws.append(list(r))

    # Sheet 2: Categories & Brands summary
    ws2 = wb.create_sheet("Categories & Brands")
    ws2.append(["Category", "Sub-Category", "Brand / Section", "Item Count"])
    # aggregate
    from collections import defaultdict
    agg = defaultdict(int)
    for cat, subcat, brand, *_ in all_rows:
        agg[(cat, subcat, brand)] += 1
    # distinct brands list
    for (cat, subcat, brand), n in sorted(agg.items()):
        ws2.append([cat, subcat, brand, n])

    # Sheet 3: Categories summary (top-level)
    ws3 = wb.create_sheet("Categories")
    ws3.append(["Category", "Sub-Category", "Item Count"])
    cat_agg = defaultdict(int)
    for cat, subcat, *_ in all_rows:
        cat_agg[(cat, subcat)] += 1
    for (cat, subcat), n in sorted(cat_agg.items()):
        ws3.append([cat, subcat, n])

    # ---- Styling ----
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F2937")
    gold_fill = PatternFill("solid", fgColor="D4AF37")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for sheet in (ws, ws2, ws3):
        # header row
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        sheet.freeze_panes = "A2"
        # body
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = left
        # zebra striping
        for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if i % 2 == 0:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="F9FAFB")

    # column widths
    widths1 = {"A": 22, "B": 26, "C": 22, "D": 34, "E": 52, "F": 16, "G": 14}
    for col, w in widths1.items():
        ws.column_dimensions[col].width = w
    for col, w in {"A": 22, "B": 26, "C": 26, "D": 12}.items():
        ws2.column_dimensions[col].width = w
    for col, w in {"A": 22, "B": 26, "C": 12}.items():
        ws3.column_dimensions[col].width = w

    # number format for price column
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")

    out = "Black_Perch_Menu.xlsx"
    wb.save(out)
    print(f"\nTotal items: {len(all_rows)}")
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
